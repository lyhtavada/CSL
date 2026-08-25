#!/usr/bin/env python3
"""
generate_program_xlsx.py — scaffold a 4-tab Joy Loyalty program proposal.

Usage:
  python3 generate_program_xlsx.py --brand "Acme Skincare" --out reports/analysis/acme-loyalty-program-2026-08-25.xlsx

Writes placeholder rows with the correct columns/structure for each tab.
Claude should fill in real numbers by editing the generated file (openpyxl)
or by re-running with --data a JSON file matching the schema in `SCHEMA` below.
"""
import argparse
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BRAND_PURPLE = "6C5CE7"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=BRAND_PURPLE, end_color=BRAND_PURPLE, fill_type="solid")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def autofit(ws, min_width=15):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = max(min_width, length + 2)


def footer(ws, row, ncols):
    ws.cell(row=row, column=1, value="Designed by Joy Loyalty").font = Font(italic=True, size=9)


def build(brand, data, out_path):
    wb = Workbook()

    # --- Tab 1: Points Program ---
    ws = wb.active
    ws.title = "Points Program"
    pv = data.get("point_valuation", {})
    ws.append(["Point Valuation"])
    ws.append(["Point value", pv.get("point_value", "1 pt = $0.01")])
    ws.append(["Earning rate", pv.get("earning_rate", "X pts per $1")])
    ws.append(["Purchases to first reward", pv.get("purchases_to_reward", "~2-3 orders")])
    ws.append(["Annual point liability estimate", pv.get("liability_estimate", "TBD")])
    ws.append([])
    row = ws.max_row + 1
    ws.append(["Earning Rules"])
    ws.append(["Rule Name", "Action", "Points Earned", "Notes/Details"])
    style_header(ws, ws.max_row, 4)
    for r in data.get("earning_rules", [["Purchase", "$1 spent", "TBD", ""]]):
        ws.append(r)
    ws.append([])
    ws.append(["Redemption Rules"])
    ws.append(["Reward Name", "Type", "Points Required", "Discount Value", "Min Order", "Notes"])
    style_header(ws, ws.max_row, 6)
    for r in data.get("redemption_rules", [["First reward", "% off", "TBD", "TBD", "TBD", ""]]):
        ws.append(r)
    footer(ws, ws.max_row + 2, 6)
    autofit(ws)

    # --- Tab 2: VIP/Membership Program ---
    ws2 = wb.create_sheet("VIP Membership")
    is_paid = data.get("paid_membership", False)
    if is_paid:
        ws2.append(["Tier Name", "Subscription Trigger", "Earning Multiplier", "Entry Reward", "Perks/Benefits"])
    else:
        ws2.append(["Tier Name", "Threshold (pts/$/orders)", "Earning Multiplier", "Entry Reward", "Perks/Benefits"])
    style_header(ws2, 1, 5)
    for r in data.get("tiers", [
        ["Tier 1", "0", "1x", "Welcome bonus", "TBD"],
        ["Tier 2", "3-5 purchases", "1.25x", "TBD", "TBD"],
        ["Tier 3", "8-12 purchases", "1.5x", "TBD", "TBD"],
        ["Tier 4 (VIP)", "15-20+ purchases", "2x", "TBD", "TBD"],
    ]):
        ws2.append(r)
    ws2.append([])
    ws2.append(["Demotion policy", data.get("demotion_policy", "TBD")])
    footer(ws2, ws2.max_row + 2, 5)
    autofit(ws2)

    # --- Tab 3: Referral Program ---
    ws3 = wb.create_sheet("Referral Program")
    ws3.append(["Element", "Configuration", "Notes"])
    style_header(ws3, 1, 3)
    for r in data.get("referral", [
        ["Referrer Reward", "TBD (≈1 purchase worth of points)", ""],
        ["Referee Reward", "10-20% off or $10-15 off", ""],
        ["Min Purchase", "At or slightly above AOV", ""],
        ["Sharing Channels", "Email, SMS, social", ""],
        ["Anti-Cheat Settings", "TBD", ""],
        ["Referral Message Template", "TBD", ""],
    ]):
        ws3.append(r)
    footer(ws3, ws3.max_row + 2, 3)
    autofit(ws3)

    # --- Tab 4: Milestones & Quest ---
    ws4 = wb.create_sheet("Milestones & Quest")
    ws4.append(["Individual Milestones"])
    ws4.append(["Milestone Name", "Type", "Target", "Reward", "Customer Message"])
    style_header(ws4, ws4.max_row, 5)
    for r in data.get("milestones", [["First order", "Order count", "1", "TBD", "TBD"]]):
        ws4.append(r)
    ws4.append([])
    ws4.append(["Quest Journey"])
    ws4.append(["Step #", "Action", "Target", "Reward", "Description"])
    style_header(ws4, ws4.max_row, 5)
    for r in data.get("quest", [[1, "TBD", "TBD", "TBD", "TBD"]]):
        ws4.append(r)
    footer(ws4, ws4.max_row + 2, 5)
    autofit(ws4)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"Saved {out_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--brand", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--data", help="path to JSON matching the schema (optional; placeholders used otherwise)")
    args = ap.parse_args()

    data = {}
    if args.data:
        with open(args.data) as f:
            data = json.load(f)

    build(args.brand, data, args.out)


if __name__ == "__main__":
    main()
