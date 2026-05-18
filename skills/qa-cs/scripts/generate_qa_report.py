#!/usr/bin/env python3
"""
Generate QA Excel report from evaluated conversations.
Usage: python3 generate_qa_report.py --input /tmp/qa_convs_<id>_<month>.json --agent <name> --month <YYYY-MM> --violations <violations_json>
Output: /Users/avada/QA_<agent>_<YYYY-MM>.xlsx
"""

import argparse
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

SEVERITY_COLORS = {
    "Urgent":   "FF4444",
    "Critical": "FF8C00",
    "High":     "FFC000",
    "Moderate": "FFE699",
    "Low":      "D9D9D9",
}

PENALTIES = {
    "Urgent": 0,     # case by case
    "Critical": 50,
    "High": 40,
    "Moderate": 30,
    "Low": 20,
}

def thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_style(ws, row, cols, bg="4A90D9"):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Path to qa_convs JSON file")
    parser.add_argument("--agent", required=True, help="Agent display name")
    parser.add_argument("--month", required=True, help="YYYY-MM")
    parser.add_argument("--violations", required=True, help="Path to violations JSON file")
    args = parser.parse_args()

    with open(args.input) as f:
        convs = json.load(f)

    with open(args.violations) as f:
        violations_data = json.load(f)
    # violations_data: list of {session_id, conv_index, violations: [{code, description, severity, penalty, evidence}]}

    # Build violation map
    viol_map = {v["session_id"]: v for v in violations_data}

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    scores = []
    all_violations = []
    for conv in convs:
        sid = conv["session_id"]
        v_entry = viol_map.get(sid, {})
        viols = v_entry.get("violations", [])
        score = 100
        for v in viols:
            score -= v.get("penalty", 0)
            all_violations.append({**v, "session_id": sid, "visitor": conv["visitor"],
                                    "date": conv["updated_at"], "url": conv["url"],
                                    "conv_index": conv["index"]})
        scores.append(score)

    avg_score = sum(scores) / len(scores) if scores else 0

    # Category breakdown
    cat_count = {}
    for v in all_violations:
        cat = v["code"][:2] if v["code"] else "?"
        cat_count[cat] = cat_count.get(cat, 0) + 1

    ws1.append(["QA Report", f"{args.agent} — {args.month}"])
    ws1.merge_cells("A1:F1")
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.append([])
    ws1.append(["Agent", args.agent])
    ws1.append(["Month", args.month])
    ws1.append(["Chats reviewed", len(convs)])
    ws1.append(["Average score", f"{avg_score:.1f}/100"])
    ws1.append(["Total violations", len(all_violations)])
    ws1.append([])
    ws1.append(["Category", "Violations"])
    header_style(ws1, ws1.max_row, 2, bg="4A90D9")
    for cat, cnt in sorted(cat_count.items()):
        labels = {"QT": "Quy trình (QT)", "KT": "Kiến thức (KT)",
                  "KN": "Kỹ năng (KN)", "BM": "Bảo mật (BM)"}
        ws1.append([labels.get(cat, cat), cnt])

    set_col_widths(ws1, [25, 20, 20, 20, 20, 20])
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Chat Details ─────────────────────────────────────────
    ws2 = wb.create_sheet("Chat Details")
    headers2 = ["#", "Visitor", "Date", "Score", "Violations", "Chat Link"]
    ws2.append(headers2)
    header_style(ws2, 1, len(headers2))

    for i, conv in enumerate(convs):
        sid = conv["session_id"]
        v_entry = viol_map.get(sid, {})
        viols = v_entry.get("violations", [])
        score = scores[i]
        viol_summary = ", ".join([v["code"] for v in viols]) if viols else "—"
        row = [conv["index"], conv["visitor"], conv["updated_at"],
               f"{score}/100", viol_summary, conv["url"]]
        ws2.append(row)

        # Color score cell
        score_cell = ws2.cell(row=i+2, column=4)
        if score < 50:
            score_cell.fill = PatternFill("solid", fgColor="FF4444")
            score_cell.font = Font(color="FFFFFF", bold=True)
        elif score < 70:
            score_cell.fill = PatternFill("solid", fgColor="FFC000")
        elif score < 90:
            score_cell.fill = PatternFill("solid", fgColor="FFE699")
        else:
            score_cell.fill = PatternFill("solid", fgColor="D9EAD3")

        for col in range(1, len(headers2)+1):
            ws2.cell(row=i+2, column=col).border = thin_border()
            ws2.cell(row=i+2, column=col).alignment = Alignment(wrap_text=True, vertical="center")

    set_col_widths(ws2, [4, 25, 16, 10, 30, 60])
    ws2.row_dimensions[1].height = 25
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Violations ───────────────────────────────────────────
    ws3 = wb.create_sheet("Violations")
    headers3 = ["#", "Chat #", "Visitor", "Date", "Code", "Description", "Severity", "Penalty", "Evidence", "Chat Link"]
    ws3.append(headers3)
    header_style(ws3, 1, len(headers3))

    for i, v in enumerate(all_violations):
        row = [
            i+1, v["conv_index"], v["visitor"], v["date"],
            v["code"], v["description"], v["severity"],
            f"-{v['penalty']}đ" if v.get("penalty") else "Case by case",
            v.get("evidence", ""), v["url"]
        ]
        ws3.append(row)

        # Color by severity
        sev = v.get("severity", "Low")
        color = SEVERITY_COLORS.get(sev, "FFFFFF")
        fill = PatternFill("solid", fgColor=color)
        for col in [7, 8]:
            ws3.cell(row=i+2, column=col).fill = fill
            ws3.cell(row=i+2, column=col).font = Font(bold=True)

        for col in range(1, len(headers3)+1):
            ws3.cell(row=i+2, column=col).border = thin_border()
            ws3.cell(row=i+2, column=col).alignment = Alignment(wrap_text=True, vertical="center")

    set_col_widths(ws3, [4, 6, 20, 16, 8, 45, 12, 12, 50, 55])
    ws3.row_dimensions[1].height = 25
    ws3.freeze_panes = "A2"

    # Save
    out_path = f"/Users/avada/QA_{args.agent}_{args.month}.xlsx"
    wb.save(out_path)
    print(f"Report saved: {out_path}")
    return out_path

if __name__ == "__main__":
    main()
